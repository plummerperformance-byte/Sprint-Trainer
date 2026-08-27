"""Load a 1080 Sprint xlsx export, parse the raw time-series, auto-detect
the actual sprint start (using the "Sprint Start" column the export marks),
build a rep dict matching ppa_service's internal shape, and POST it into
the running service via /api/c/dev/load_rep.

Usage:
    python load_1080_xlsx.py "C:\\path\\to\\Tyrone 1080.xlsx"

Run ppa_service.py (or ppa_app.py) first so the endpoint is available.
"""
from __future__ import annotations

import json
import math
import sys
import urllib.request
from pathlib import Path
from typing import Optional

import openpyxl

SERVICE_URL = "http://127.0.0.1:8765"
CHART_SAMPLE_BUDGET = 600  # decimate raw 1 kHz → ~600 chart points (≈50 Hz effective)
SPLIT_LENGTH_M = 5.0       # 1080-style uniform split bucket


def parse_xlsx(path: Path, start_foot: str = "left") -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    # Find raw-data sheet (one with "Raw Data" in the name)
    raw_name = next((s for s in wb.sheetnames if "raw data" in s.lower()), None)
    if not raw_name:
        raise ValueError(f"no Raw Data sheet found in {path.name}; sheets={wb.sheetnames}")
    ws = wb[raw_name]

    # Header row is row 1. Columns we care about:
    #   A time(ms)  B load(g)  C speed(mm/s)  D position(mm)  E Sprint Start
    # Find Sprint Start marker — it's a single number on row 2 indicating the
    # 0-indexed sample where the actual sprint kicks off.
    sprint_start_idx = None
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if rows and len(rows[0]) >= 5 and isinstance(rows[0][4], (int, float)):
        sprint_start_idx = int(rows[0][4])
    print(f"  raw rows: {len(rows)}  sprint start sample: {sprint_start_idx}")
    if sprint_start_idx is None or sprint_start_idx < 0:
        raise ValueError("Sprint Start column missing or empty in raw-data sheet")

    # Slice from sprint start onward; subtract start position so it's rep-relative.
    run_rows = rows[sprint_start_idx:]
    if not run_rows:
        raise ValueError("Sprint Start index is past the end of the raw data")
    t0 = run_rows[0][0]
    pos0_mm = run_rows[0][3]

    # Compute aggregates from FULL 1 kHz data. Decimate samples just for chart.
    # body mass for power-rel calculation (Dashboard sheet, "Weight (kg)" cell)
    # + athlete name + Morin pre-computes (F0, V0, Pmax, Slope) if 1080 wrote them
    body_mass_kg = None
    athlete_name = None
    morin = {"f0_n": None, "v0_mps": None, "pmax_w": None,
             "f0_rel_nkg": None, "pmax_rel_wkg": None, "fv_slope_per_kg": None}
    if "Dashboard" in wb.sheetnames:
        for row in wb["Dashboard"].iter_rows(values_only=True):
            if not row or row[0] is None: continue
            label = str(row[0]).strip().lower()
            val = row[1]
            try: f = float(val) if val is not None else None
            except Exception: f = None
            if label.startswith("weight") and f is not None: body_mass_kg = f
            elif label == "athlete" and val is not None: athlete_name = str(val)
            elif label.startswith("f₀ / kg") and f is not None: morin["f0_rel_nkg"] = f
            elif label.startswith("f₀") and f is not None: morin["f0_n"] = f
            elif label.startswith("v₀") and f is not None: morin["v0_mps"] = f
            elif label.startswith("pmax / kg") and f is not None: morin["pmax_rel_wkg"] = f
            elif label.startswith("pmax") and f is not None: morin["pmax_w"] = f
            elif label.startswith("slope") and f is not None and body_mass_kg:
                # Dashboard reports slope in N.s/m (not per kg); normalise to per-kg
                # which is what the Morin/Lahti cutoffs operate on.
                morin["fv_slope_per_kg"] = round(f / body_mass_kg, 4)

    # Build full-rate metric arrays
    times_s = []
    speeds_mps = []
    forces_n = []
    pos_m = []
    horiz_forces_n = []   # col 13 — horizontal force (Morin's F_h)
    rf_pct_list = []      # col 18 — RF% (horizontal / resultant)
    for r in run_rows:
        if r[0] is None or r[2] is None or r[3] is None:
            continue
        t_ms = r[0] - t0
        v_mps = (r[2] or 0) / 1000.0
        load_g = r[1] or 0
        f_n = (load_g / 1000.0) * 9.81
        p_mm = (r[3] or 0) - pos0_mm
        times_s.append(t_ms / 1000.0)
        speeds_mps.append(v_mps)
        forces_n.append(f_n)
        pos_m.append(p_mm / 1000.0)
        # Horizontal force + RF% if present (cols 13 / 18, 0-indexed 12 / 17)
        horiz_forces_n.append(r[12] if len(r) > 12 else None)
        rf = r[17] if len(r) > 17 else None
        # 1080 stores RF as 0–1 fraction; promote to %
        rf_pct_list.append(float(rf) * 100 if rf is not None else None)

    if len(times_s) < 2:
        raise ValueError("not enough samples in the run section")

    # Aggregates (full-rate)
    peak_v = max(speeds_mps)
    peak_v_idx = speeds_mps.index(peak_v)
    peak_f = max(forces_n)
    peak_f_idx = forces_n.index(peak_f)
    powers_w = [f * v for f, v in zip(forces_n, speeds_mps)]
    peak_p = max(powers_w)

    n = len(speeds_mps)
    avg_v = sum(speeds_mps) / n
    avg_f = sum(forces_n) / n
    avg_p = sum(powers_w) / n

    # Acceleration (Δv / Δt)
    accs = []
    for i in range(1, n):
        dt = times_s[i] - times_s[i - 1]
        if dt > 0:
            accs.append((speeds_mps[i] - speeds_mps[i - 1]) / dt)
    peak_a = max(accs) if accs else 0.0
    avg_a = sum(accs) / len(accs) if accs else 0.0

    # ∫F·v dt  and  ∫F dt
    work_j = 0.0
    impulse_ns = 0.0
    for i in range(1, n):
        dt = times_s[i] - times_s[i - 1]
        if dt <= 0: continue
        f_avg = (forces_n[i] + forces_n[i - 1]) / 2
        v_avg = (speeds_mps[i] + speeds_mps[i - 1]) / 2
        work_j += f_avg * v_avg * dt
        impulse_ns += f_avg * dt

    duration_s = times_s[-1]
    max_extension_m = max(pos_m)

    # Splits (5/10/20 m — kept for back-compat with existing UI; extended below)
    splits = {}
    for marker in (5, 10, 20):
        for i, p in enumerate(pos_m):
            if p >= marker:
                splits[str(marker)] = round(times_s[i], 3)
                break

    # Sprint phase segmentation (90% / 95% of peak v)
    accel_end_ms = None
    decel_start_ms = None
    past_peak = False
    for i, v in enumerate(speeds_mps):
        if accel_end_ms is None and v >= 0.90 * peak_v:
            accel_end_ms = int(times_s[i] * 1000)
        if v >= peak_v:
            past_peak = True
        if past_peak and decel_start_ms is None and v < 0.95 * peak_v:
            decel_start_ms = int(times_s[i] * 1000)

    # ---- 1080-App sprint metric extensions ----
    # Time / distance to peak v + 90% peak v (acceleration profile shape)
    time_to_max_v_s = times_s[peak_v_idx]
    dist_to_max_v_m = pos_m[peak_v_idx]
    time_to_90pct_v_s = None
    dist_to_90pct_v_m = None
    target_90 = 0.90 * peak_v
    for i, v in enumerate(speeds_mps):
        if v >= target_90:
            time_to_90pct_v_s = times_s[i]
            dist_to_90pct_v_m = pos_m[i]
            break
    end_v = speeds_mps[-1]
    v_dropoff_pct = (peak_v - end_v) / peak_v * 100 if peak_v > 0 else 0.0

    # Splits extended to 5/10/20/30/40 m
    splits_extended = {}
    for marker in (5, 10, 20, 30, 40):
        for i, p in enumerate(pos_m):
            if p >= marker:
                splits_extended[str(marker)] = round(times_s[i], 3)
                break

    # ---- Step detection + foot labels + L/R asymmetry ----
    # Primary detector = speed residual (doc §2); it beat the cable-force detector
    # decisively in the A/B on real 1080 data (speed found ~2x the steps at the
    # right rhythm). The force detector stays available for heavy-resisted sprints
    # where cable tension may carry a cleaner per-step signal — worth re-checking.
    step_events = detect_steps_speed_residual(times_s, speeds_mps, pos_m)
    step_events = annotate_steps(step_events, times_s, forces_n)
    step_events = label_feet(step_events, start_foot)
    step_aggs = compute_step_aggregates(step_events)
    asymmetry = compute_asymmetry(step_events)

    # ---- RFmax + Drf are NOT derivable from cable-only sprint data ----
    # Morin's RFmax = horizontal_GRF / resultant_GRF, requires either a
    # force plate or Samozino's inverse-dynamics method on a *free* sprint.
    # Cable load during a resisted sprint is not horizontal GRF — it's a
    # small fraction of total propulsion. The 1080 xlsx Force/RF% columns
    # (cols 13/18) are populated only in the pre-sprint loading rows, not
    # during the run, and reflect cable-loading mechanics rather than
    # sprint mechanics. Mechanical-effectiveness module stays silent for
    # cable-only imports until we add a force plate or implement the
    # full Samozino model on a free-sprint capture path.
    rfmax_pct = None
    drf = None

    # 1080-shaped split report (uniform 5 m buckets)
    split_report = build_split_report(times_s, speeds_mps, forces_n, powers_w, accs, pos_m, SPLIT_LENGTH_M)

    # Time-to-peak metrics (ms)
    ttpf_ms = int(times_s[peak_f_idx] * 1000)
    ttps_ms = int(times_s[peak_v_idx] * 1000)

    # Decimate to ~CHART_SAMPLE_BUDGET points for the UI chart
    step = max(1, n // CHART_SAMPLE_BUDGET)
    samples = []
    for i in range(0, n, step):
        a = accs[i - 1] if 0 < i <= len(accs) else 0.0
        samples.append({
            "t_ms": int(times_s[i] * 1000),
            "v_mps": round(speeds_mps[i], 3),
            "F_N": round(forces_n[i], 1),
            "P_W": round(powers_w[i], 1),
            "a_mps2": round(a, 3),
            "pos_m": round(pos_m[i], 3),
        })
    print(f"  decimated samples: {len(samples)} (every {step}th of {n})")

    # Per-kg power if body mass known
    peak_power_rel_wkg = round(peak_p / body_mass_kg, 2) if body_mass_kg else None

    rep = {
        "rep_idx": 1,
        "duration_s": round(duration_s, 3),
        "max_extension_m": round(max_extension_m, 3),
        "total_distance_m": round(max_extension_m, 3),
        "total_time_s": round(duration_s, 3),
        "peak_speed_mps": round(peak_v, 3),
        "top_speed_mps": round(peak_v, 3),
        "peak_force_n": round(peak_f, 1),
        "peak_power_w": round(peak_p, 1),
        "peak_power_rel_wkg": peak_power_rel_wkg,
        "peak_acceleration_mps2": round(peak_a, 3),
        "peak_torque_pct": None,  # not present in xlsx (1080 doesn't expose torque %)
        "avg_speed_mps": round(avg_v, 3),
        "avg_force_n": round(avg_f, 1),
        "avg_power_w": round(avg_p, 1),
        "avg_acceleration_mps2": round(avg_a, 3),
        "work_j": round(work_j, 1),
        "impulse_ns": round(impulse_ns, 1),
        "ttpf_ms": ttpf_ms,
        "ttps_ms": ttps_ms,
        "accel_end_ms": accel_end_ms,
        "decel_start_ms": decel_start_ms,
        "is_eccentric": False,
        "splits_s": splits,
        "splits_s_extended": splits_extended,
        "split_report": split_report,
        # Morin mechanical effectiveness — computed from cols 13/18 of the xlsx
        "rf_max_pct": rfmax_pct,
        "drf": drf,
        # 1080-App sprint_metrics extensions
        "time_to_max_v_s": round(time_to_max_v_s, 3),
        "dist_to_max_v_m": round(dist_to_max_v_m, 2),
        "time_to_90pct_v_s": round(time_to_90pct_v_s, 3) if time_to_90pct_v_s is not None else None,
        "dist_to_90pct_v_m": round(dist_to_90pct_v_m, 2) if dist_to_90pct_v_m is not None else None,
        "v_dropoff_pct": round(v_dropoff_pct, 1),
        # Step events + aggregates (1080-App step_events table shape)
        "step_events": step_events,
        **step_aggs,
        # Foot-labelled L/R asymmetry (declared start foot; doc §4/§5)
        "asymmetry": asymmetry,
        "samples": samples,
        # Morin sprint-FV outputs (1080's own per-rep regression, imported from
        # Dashboard sheet). Keys mirror the 1080-App `sprint_metrics` table.
        "f0_n": morin["f0_n"],
        "f0_rel_nkg": morin["f0_rel_nkg"],
        "v0_mps": morin["v0_mps"],
        "pmax_w_morin": morin["pmax_w"],   # 1080's Pmax (multi-load curve), not raw F.v peak
        "pmax_rel_wkg": morin["pmax_rel_wkg"] if morin["pmax_rel_wkg"] is not None
                        else (round(peak_p / body_mass_kg, 2) if body_mass_kg else None),
        "fv_slope_per_kg": morin["fv_slope_per_kg"],
        "_meta": {
            "source": "1080 xlsx",
            "athlete_name": athlete_name,
            "body_mass_kg": body_mass_kg,
            "raw_sample_count": n,
            "sprint_start_index": sprint_start_idx,
        },
    }
    return rep


def _smooth(xs, window=15):
    """Simple centred moving-average smooth. Window is in samples (~15 ms at 1 kHz).
    Reduces high-frequency noise without smearing real peaks (~50-100 ms wide)."""
    half = window // 2
    n = len(xs)
    out = []
    for i in range(n):
        a = max(0, i - half); b = min(n, i + half + 1)
        out.append(sum(xs[a:b]) / (b - a))
    return out


# Physiological sanity window for a single sprint step (handoff doc §2 / §7).
# These bound the accepted inter-event gap. Values outside are FLAGGED, not
# silently dropped — a long gap usually means a *missed* peak that a reviewer
# should see, per the doc's confidence/manual-correction guidance (§8).
STEP_MIN_S = 0.11      # < ~9 Hz between strikes is implausibly fast
STEP_MAX_S = 0.45      # > this gap ≈ a missed step
STEP_MIN_LEN_M = 0.25  # a "step" shorter than this is almost certainly noise
STEP_MAX_LEN_M = 3.5   # longer than this is implausible for one step


def detect_steps(t_s, f_n, pos_m, min_period_s=0.15, min_prominence_n=4.0, smooth_window=15,
                 min_step_len_m=STEP_MIN_LEN_M):
    """Peak-detect foot strikes on the cable force trace.

    Cable trainers maintain constant baseline tension so peaks are subtle
    (typically 5-30 N above the local valley, not 0 → peak like a force
    plate). Use prominence-based detection on a smoothed trace:
      1. Light moving-average smooth (~15 ms) to remove digitisation noise
      2. Find every local maximum
      3. Compute prominence (how far it rises above the lower of the two
         adjacent valleys); discard peaks below `min_prominence_n`
      4. Reject a candidate that sits too close to the previous accepted event
         in TIME (`min_period_s`) OR DISTANCE (`min_step_len_m`) — keep the more
         prominent of the pair. Distance-dedup (doc §2) catches one-step splits
         the time gate alone misses at high cable speed.
      5. Tag each accepted interval whose period/length falls outside the
         physiological window with a `flags` list (doc §7 validation), rather
         than dropping it — a long gap usually means a *missed* step to review.

    Returns list of step dicts (1080-App step_events shape), each with `flags`.
    """
    if len(f_n) < 3:
        return []
    fs = _smooth(f_n, smooth_window)

    # Find ALL local maxima (no threshold yet)
    maxima = []
    for i in range(1, len(fs) - 1):
        if fs[i] > fs[i - 1] and fs[i] >= fs[i + 1]:
            maxima.append(i)
    if not maxima:
        return []

    # For each maximum, compute prominence vs adjacent valleys (look in a
    # ±0.5 s window, which comfortably brackets one full step at 2-5 Hz).
    win = max(20, int(0.5 / max(t_s[1] - t_s[0], 1e-4)))  # samples in 0.5 s
    accepted = []
    for idx in maxima:
        a = max(0, idx - win); b = min(len(fs), idx + win + 1)
        valley_left = min(fs[a:idx + 1]) if idx > a else fs[idx]
        valley_right = min(fs[idx:b]) if idx < b - 1 else fs[idx]
        prominence = fs[idx] - max(valley_left, valley_right)
        if prominence < min_prominence_n:
            continue
        if accepted:
            prev = accepted[-1]
            too_close = ((t_s[idx] - t_s[prev]) < min_period_s
                         or (pos_m[idx] - pos_m[prev]) < min_step_len_m)
            if too_close:
                # same step split into two candidates — keep the more prominent
                if fs[idx] > fs[prev]:
                    accepted[-1] = idx
                continue
        accepted.append(idx)

    # Build step events using ORIGINAL (un-smoothed) force values for the recorded peak
    steps = []
    for i, idx in enumerate(accepted):
        prev_idx = accepted[i - 1] if i > 0 else None
        period_ms = round((t_s[idx] - t_s[prev_idx]) * 1000, 1) if prev_idx is not None else None
        length_m = round(pos_m[idx] - pos_m[prev_idx], 3) if prev_idx is not None else None
        inst_freq_hz = round(1000 / period_ms, 2) if period_ms and period_ms > 0 else None
        step_speed = (round(length_m / (period_ms / 1000.0), 3)
                      if (length_m is not None and period_ms) else None)
        flags = []
        if period_ms is not None:
            if period_ms > STEP_MAX_S * 1000: flags.append("long_gap")    # likely missed step
            elif period_ms < STEP_MIN_S * 1000: flags.append("short_gap")  # likely double
        if length_m is not None:
            if length_m > STEP_MAX_LEN_M: flags.append("long_step")
            elif length_m < STEP_MIN_LEN_M: flags.append("short_step")
        steps.append({
            "step_number": i + 1,
            "t_strike_s": round(t_s[idx], 4),
            "pos_m": round(pos_m[idx], 3),
            "peak_force_n": round(f_n[idx], 1),
            "step_period_ms": period_ms,
            "step_length_m": length_m,
            "step_frequency_hz": inst_freq_hz,
            "step_speed_mps": step_speed,
            "flags": flags,
        })
    return steps


def _rolling_med_mad(sig, w):
    """Rolling median + MAD (median absolute deviation) over a w-sample window.
    Basis for the adaptive step-accept threshold (handoff §5.1)."""
    h = w // 2
    n = len(sig)
    med = [0.0] * n
    mad = [0.0] * n
    for i in range(n):
        a = max(0, i - h); b = min(n, i + h + 1)
        win = sorted(sig[a:b]); m = win[len(win) // 2]
        med[i] = m
        mad[i] = sorted(abs(x - m) for x in win)[len(win) // 2]
    return med, mad


def _step_flags(period_ms, length_m):
    """Physiological-window validation flags shared by both detectors (doc §7)."""
    flags = []
    if period_ms is not None:
        if period_ms > STEP_MAX_S * 1000: flags.append("long_gap")
        elif period_ms < STEP_MIN_S * 1000: flags.append("short_gap")
    if length_m is not None:
        if length_m > STEP_MAX_LEN_M: flags.append("long_step")
        elif length_m < STEP_MIN_LEN_M: flags.append("short_step")
    return flags


def _recover_missed_strikes(accepted, t_s, sig, min_period_s, gap_split_ratio=1.6):
    """Split over-long steps that swallowed a missed strike.

    A single missed strike doubles a step's length and period — e.g. Tyrone's
    end-of-run 3.98 m 'step', where the real foot strike fell ~0.01 below the
    adaptive accept threshold during deceleration and got merged into its
    neighbour. Loosening the global threshold to catch it spawns false positives
    elsewhere; instead, for every gap longer than `gap_split_ratio` x the LOCAL
    median step period, estimate how many strikes are missing
    (round(gap / local_median) - 1) and re-insert that many at the tallest
    sub-threshold residual maxima inside the gap. When the signal offers no clear
    peak (a true data drop-out), fall back to even time subdivision. Returns
    (new_accepted, recovered_index_set) so callers can flag the inserted strikes
    as lower-confidence."""
    if len(accepted) < 4:
        return accepted, set()
    gaps = sorted(t_s[accepted[i]] - t_s[accepted[i - 1]] for i in range(1, len(accepted)))
    global_med = gaps[len(gaps) // 2] or min_period_s

    def local_med(k):
        lo = max(1, k - 3); hi = min(len(accepted), k + 4)
        w = sorted(t_s[accepted[j]] - t_s[accepted[j - 1]] for j in range(lo, hi))
        return (w[len(w) // 2] if w else global_med) or global_med

    recovered = set()
    out = [accepted[0]]
    for k in range(1, len(accepted)):
        a, b = accepted[k - 1], accepted[k]
        gap = t_s[b] - t_s[a]
        med = local_med(k)
        n_expected = int(round(gap / med)) if med > 0 else 1
        if gap > gap_split_ratio * med and n_expected >= 2:
            n_missing = n_expected - 1
            cands = sorted((i for i in range(a + 1, b)
                            if sig[i] > sig[i - 1] and sig[i] >= sig[i + 1]),
                           key=lambda i: sig[i], reverse=True)
            picks = []
            for i in cands:
                if (t_s[i] - t_s[a]) < min_period_s or (t_s[b] - t_s[i]) < min_period_s:
                    continue
                if any(abs(t_s[i] - t_s[p]) < min_period_s for p in picks):
                    continue
                picks.append(i)
                if len(picks) >= n_missing:
                    break
            if len(picks) < n_missing:                 # data drop-out: subdivide evenly
                for m in range(1, n_expected):
                    tt = t_s[a] + gap * m / n_expected
                    j = min(range(a + 1, b), key=lambda i: abs(t_s[i] - tt))
                    if j not in picks:
                        picks.append(j)
            for i in sorted(picks)[:n_missing]:
                recovered.add(i)
                out.append(i)
        out.append(b)
    return out, recovered


# Speed-residual defaults validated by A/B against the force detector on the
# Tyrone 1080 workbook (short 25 ms / long 200 ms / no prominence floor gave 23
# clean detections + 3 flagged gaps ≈ 26 true steps, matching the doc's 25-26 /
# 4.39 Hz reference; the force detector found only 13 at 2.2 Hz). A long window
# past ~300 ms starts spanning whole steps and cancels the step signal.
def detect_steps_speed_residual(t_s, v_mps, pos_m, short_ms=25, long_ms=200,
                                min_period_s=0.11, min_step_len_m=STEP_MIN_LEN_M,
                                adaptive_k=0.2, thr_win_ms=400, min_prominence_mps=None,
                                gap_split_ratio=1.6):
    """Doc §2 speed-residual step detector (the doc's recommended first pass).

    Removes the broad acceleration trend from the tether speed and keeps the
    rhythmic step-scale pulses:
        v_short = MA(speed, ~25-40 ms)   (short_ms)
        v_long  = MA(speed, ~200-300 ms) (long_ms)
        step_signal = v_short - v_long
    Local maxima of step_signal are candidate strikes; a candidate too close to
    the previous accepted event in TIME or DISTANCE is merged (keep the taller
    residual peak). Same event shape + `flags` as the force detector so the two
    are directly comparable.
    """
    n = len(v_mps)
    if n < 5:
        return []
    dt = max(t_s[1] - t_s[0], 1e-4)
    fs_hz = 1.0 / dt
    w_short = max(1, int(round(short_ms / 1000.0 * fs_hz)))
    w_long = max(w_short + 1, int(round(long_ms / 1000.0 * fs_hz)))
    v_short = _smooth(v_mps, w_short)
    v_long = _smooth(v_mps, w_long)
    sig = [a - b for a, b in zip(v_short, v_long)]

    # Adaptive accept threshold (handoff §5.1): rolling median + k·MAD tracks the
    # local noise floor, recovering the steps a fixed floor misses during
    # acceleration. On the Tyrone workbook this lifts detection from 23 @ 3.73 Hz
    # (3 missed) to 26 @ 4.24 Hz (0 missed) — matching the 25-26 / 4.39 Hz reference.
    if adaptive_k is not None:
        tw = max(5, int(round(thr_win_ms / 1000.0 * fs_hz)))
        med, mad = _rolling_med_mad(sig, tw)
        thr = [med[i] + adaptive_k * 1.4826 * mad[i] for i in range(n)]
    else:
        fl = min_prominence_mps or 0.0
        thr = [fl] * n
    maxima = [i for i in range(1, n - 1)
              if sig[i] > sig[i - 1] and sig[i] >= sig[i + 1] and sig[i] > thr[i]]
    if not maxima:
        return []
    accepted = []
    for idx in maxima:
        if accepted:
            prev = accepted[-1]
            if (t_s[idx] - t_s[prev]) < min_period_s or (pos_m[idx] - pos_m[prev]) < min_step_len_m:
                if sig[idx] > sig[prev]:
                    accepted[-1] = idx
                continue
        accepted.append(idx)

    # Recover strikes the accept threshold missed inside an over-long gap, so a
    # single miss no longer reads as one impossible 3.9 m step (see helper).
    if gap_split_ratio:
        accepted, recovered = _recover_missed_strikes(
            accepted, t_s, sig, min_period_s, gap_split_ratio)
    else:
        recovered = set()

    steps = []
    for i, idx in enumerate(accepted):
        prev_idx = accepted[i - 1] if i > 0 else None
        period_ms = round((t_s[idx] - t_s[prev_idx]) * 1000, 1) if prev_idx is not None else None
        length_m = round(pos_m[idx] - pos_m[prev_idx], 3) if prev_idx is not None else None
        inst_freq = round(1000 / period_ms, 2) if period_ms and period_ms > 0 else None
        step_speed = (round(length_m / (period_ms / 1000.0), 3)
                      if (length_m is not None and period_ms) else None)
        flags = _step_flags(period_ms, length_m)
        if idx in recovered:
            flags.append("recovered")
        steps.append({
            "step_number": i + 1,
            "t_strike_s": round(t_s[idx], 4),
            "pos_m": round(pos_m[idx], 3),
            "peak_speed_mps": round(v_mps[idx], 3),
            "step_period_ms": period_ms,
            "step_length_m": length_m,
            "step_frequency_hz": inst_freq,
            "step_speed_mps": step_speed,
            "flags": flags,
        })
    return steps


def annotate_steps(step_events, t_s, f_n):
    """Add the two per-step columns the 1080 table shows that the speed-residual
    detector doesn't produce itself: same-foot stride length and peak cable force."""
    import bisect
    for i, e in enumerate(step_events):
        # stride = same foot to same foot = this step + the previous one
        e["stride_length_m"] = round(e["pos_m"] - step_events[i - 2]["pos_m"], 3) if i >= 2 else None
        # peak cable force within this step's interval
        if i == 0:
            e["peak_force_n"] = None
        else:
            lo = bisect.bisect_left(t_s, step_events[i - 1]["t_strike_s"])
            hi = bisect.bisect_right(t_s, e["t_strike_s"])
            seg = f_n[lo:hi]
            e["peak_force_n"] = round(max(seg), 1) if seg else None
    return step_events


def compute_step_aggregates(steps):
    """Roll up a step-events list into per-rep aggregates."""
    if len(steps) < 2:
        return {"total_steps": len(steps)}
    valid = [s for s in steps if s.get("step_period_ms") is not None]
    if not valid:
        return {"total_steps": len(steps)}
    periods = [s["step_period_ms"] for s in valid]
    lengths = [s["step_length_m"] for s in valid if s.get("step_length_m") is not None]
    avg_period = sum(periods) / len(periods)
    step_freq_hz = 1000 / avg_period if avg_period > 0 else 0
    avg_length = sum(lengths) / len(lengths) if lengths else 0
    var_l = sum((l - avg_length) ** 2 for l in lengths) / len(lengths) if lengths else 0
    std_length = var_l ** 0.5
    flagged = sum(1 for s in steps if s.get("flags"))
    return {
        "total_steps": len(steps),
        "step_freq_hz": round(step_freq_hz, 2),
        "avg_step_length_m": round(avg_length, 3),
        "step_length_std_m": round(std_length, 3),
        "avg_step_period_ms": round(avg_period, 1),
        "flagged_steps": flagged,
        "step_confidence": round(1 - flagged / len(steps), 2),
    }


def label_feet(steps, start_foot="left"):
    """Alternate L/R foot labels from a user-declared start foot (doc §4).

    Tether data cannot know true foot identity (doc §8) — this is a *declared*
    start alternation, not a measurement. A step preceded by a 'long_gap' flag
    (a likely missed strike) flips the true alternation from that point, so it
    is marked `foot_suspect` for a reviewer to relabel from there.
    """
    start = "left" if str(start_foot).lower().startswith("l") else "right"
    other = "right" if start == "left" else "left"
    for i, s in enumerate(steps):
        s["foot"] = start if i % 2 == 0 else other
        s["foot_suspect"] = "long_gap" in (s.get("flags") or [])
    return steps


def _asym_pct(left, right):
    """Normalised symmetry index: 200*(L-R)/(L+R). +ve = left larger (doc §5)."""
    if left is None or right is None:
        return None
    denom = left + right
    return round(200.0 * (left - right) / denom, 1) if denom else None


def compute_asymmetry(steps, zone_len_m=10.0):
    """Left/right asymmetry across the step series (handoff doc §5).

    Reports three views for each spatio-temporal metric:
      * global  — whole-sprint L-mean vs R-mean (marked low-confidence: rising
                  speed through the sprint biases it, doc §5/§8)
      * pairwise — mean of consecutive L/R pair asymmetries (accel-robust)
      * zones    — asymmetry within 0-10 / 10-20 / … m distance buckets
    Returns None if fewer than two labelled steps.
    """
    # Only clean steps feed the L/R comparison. A flagged step — above all a
    # `recovered` (interpolated) strike — has an approximate position/timing, so
    # including it would fabricate a metre-scale left-vs-right difference that
    # isn't real. Aggregates (avg length etc.) still use every step; asymmetry
    # is deliberately stricter.
    labelled = [s for s in steps
                if s.get("foot") and s.get("step_period_ms") is not None
                and not s.get("flags")]
    if len(labelled) < 2:
        return None
    METRICS = ("step_length_m", "step_period_ms", "step_frequency_hz",
               "step_speed_mps", "peak_force_n")

    def mean(vals):
        vals = [v for v in vals if v is not None]
        return sum(vals) / len(vals) if vals else None

    lefts = [s for s in labelled if s["foot"] == "left"]
    rights = [s for s in labelled if s["foot"] == "right"]

    glob = {k: _asym_pct(mean([s.get(k) for s in lefts]),
                         mean([s.get(k) for s in rights])) for k in METRICS}

    pairwise = {k: [] for k in METRICS}
    for a, b in zip(labelled, labelled[1:]):
        if a["foot"] == b["foot"]:
            continue  # a missed step broke the alternation — skip this pair
        lft, rgt = (a, b) if a["foot"] == "left" else (b, a)
        for k in METRICS:
            p = _asym_pct(lft.get(k), rgt.get(k))
            if p is not None:
                pairwise[k].append(p)
    pairwise_mean = {k: (round(sum(v) / len(v), 1) if v else None)
                     for k, v in pairwise.items()}

    max_pos = max(s["pos_m"] for s in labelled)
    zones = []
    for z in range(int(max_pos // zone_len_m) + 1):
        lo, hi = z * zone_len_m, (z + 1) * zone_len_m
        zl = [s for s in lefts if lo <= s["pos_m"] < hi]
        zr = [s for s in rights if lo <= s["pos_m"] < hi]
        if not zl or not zr:
            continue
        row = {"zone": f"{int(lo)}-{int(hi)}m", "left_steps": len(zl), "right_steps": len(zr)}
        for k in METRICS:
            row[k] = _asym_pct(mean([s.get(k) for s in zl]), mean([s.get(k) for s in zr]))
        zones.append(row)

    # --- stride (same-foot) asymmetry: the trustworthy view ---
    # A single step is the wrong unit for L/R comparison: the detector splits each
    # gait cycle into two steps at a slightly uneven point, so consecutive steps
    # alternate long/short even when the athlete is symmetric (Tyrone's ~2.0/1.9 m
    # split of a symmetric ~4.4 m stride). Comparing SAME-FOOT strides (L->L vs
    # R->R) sidesteps that entirely, and pairing each left stride with the
    # overlapping right stride cancels the acceleration trend (both cover almost
    # the same ground). This is what makes the number match reality instead of the
    # ~20% the raw step comparison invents.
    def _foot_strides(foot):
        seq = [s for s in labelled if s["foot"] == foot]
        return [(seq[i]["pos_m"] - seq[i - 1]["pos_m"],                      # stride length m
                 (seq[i]["t_strike_s"] - seq[i - 1]["t_strike_s"]) * 1000.0)  # stride period ms
                for i in range(1, len(seq))]

    Ls, Rs = _foot_strides("left"), _foot_strides("right")
    stride = {}
    if len(Ls) >= 2 and len(Rs) >= 2:
        npair = min(len(Ls), len(Rs))
        len_p = [_asym_pct(Ls[i][0], Rs[i][0]) for i in range(npair)]
        per_p = [_asym_pct(Ls[i][1], Rs[i][1]) for i in range(npair)]
        len_p = [p for p in len_p if p is not None]
        per_p = [p for p in per_p if p is not None]
        lmean = sum(Ls[i][0] for i in range(npair)) / npair
        rmean = sum(Rs[i][0] for i in range(npair)) / npair
        stride = {
            "n_pairs": npair,
            "step_length_m": round(sum(len_p) / len(len_p), 1) if len_p else None,
            "step_period_ms": round(sum(per_p) / len(per_p), 1) if per_p else None,
            "len_diff_m": round(lmean - rmean, 3),   # L - R, native metres
            "left_stride_m": round(lmean, 3),
            "right_stride_m": round(rmean, 3),
        }

    return {
        "start_foot": (steps[0].get("foot") if steps else labelled[0]["foot"]),
        "declared": True,
        "global": glob,
        "global_confidence": "low (acceleration-biased)",
        "pairwise_mean": pairwise_mean,
        "zones": zones,
        "stride": stride,
        "stride_confidence": "same-foot stride, acceleration-controlled (trustworthy)",
    }


def build_split_report(t_s, v_mps, f_n, p_w, accs, pos_m, split_length_m):
    if not pos_m:
        return {"splits": [], "splitLength": split_length_m, "isYards": False}
    max_pos = max(pos_m)
    n_splits = int(max_pos // split_length_m)
    if n_splits <= 0:
        return {"splits": [], "splitLength": split_length_m, "isYards": False}
    splits_out = []
    for i in range(n_splits):
        start_m = i * split_length_m
        end_m = (i + 1) * split_length_m
        idxs = [j for j, p in enumerate(pos_m) if start_m <= p < end_m]
        if not idxs:
            continue
        n = len(idxs)
        bs = [v_mps[j] for j in idxs]
        bf = [f_n[j] for j in idxs]
        bp = [p_w[j] for j in idxs]
        ba = [accs[max(0, j - 1)] if j > 0 else 0.0 for j in idxs]
        splits_out.append({
            "start": round(start_m, 2),
            "end": round(end_m, 2),
            "time": round(t_s[idxs[-1]] - t_s[idxs[0]], 3),
            "averages": {
                "avg_speed": round(sum(bs) / n, 3),
                "avg_force": round(sum(bf) / n, 1),
                "avg_power": round(sum(bp) / n, 1),
                "avg_acceleration": round(sum(ba) / n, 3),
            },
            "peaks": {
                "peak_speed": round(max(bs), 3),
                "peak_force": round(max(bf), 1),
                "peak_power": round(max(bp), 1),
                "peak_acceleration": round(max(ba), 3),
            },
        })
    return {"splits": splits_out, "splitLength": split_length_m, "isYards": False}


def find_or_create_athlete(name: str) -> Optional[int]:
    """Look up athlete by name (exact case-insensitive) or create one. Returns id."""
    if not name:
        return None
    try:
        with urllib.request.urlopen(SERVICE_URL + "/api/athletes", timeout=5) as r:
            athletes = json.loads(r.read().decode("utf-8"))
        for a in athletes:
            if a.get("name", "").strip().lower() == name.strip().lower():
                return a["id"]
        # Not found — create
        body = json.dumps({"name": name}).encode("utf-8")
        req = urllib.request.Request(
            SERVICE_URL + "/api/athletes",
            data=body, headers={"Content-Type": "application/json"}, method="POST",
        )
        with urllib.request.urlopen(req, timeout=5) as r:
            j = json.loads(r.read().decode("utf-8"))
            print(f"  created athlete '{name}' id={j['id']}")
            return j["id"]
    except Exception as e:
        print(f"  athlete lookup/create failed: {e}", file=sys.stderr)
        return None


def post_rep(rep: dict, athlete_id: Optional[int] = None) -> None:
    payload = {"reps": [rep], "athlete_id": athlete_id} if athlete_id else rep
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        SERVICE_URL + "/api/c/dev/load_rep",
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        print(f"  POST -> {r.status}: {r.read().decode('utf-8')}")


def main(path_str: str, start_foot: str = "left") -> None:
    path = Path(path_str)
    if not path.exists():
        print(f"file not found: {path}", file=sys.stderr); sys.exit(1)
    print(f"loading {path.name}  (start foot: {start_foot})")
    rep = parse_xlsx(path, start_foot)
    print(f"  athlete:         {rep['_meta'].get('athlete_name')} ({rep['_meta'].get('body_mass_kg')} kg)")
    print(f"  duration:        {rep['duration_s']:.2f} s")
    print(f"  distance:        {rep['max_extension_m']:.2f} m")
    print(f"  peak speed:      {rep['peak_speed_mps']:.2f} m/s @ {rep['time_to_max_v_s']:.2f}s / {rep['dist_to_max_v_m']:.1f}m")
    if rep.get('time_to_90pct_v_s') is not None:
        print(f"  90% peak v:      {rep['time_to_90pct_v_s']:.2f}s / {rep['dist_to_90pct_v_m']:.1f}m")
    print(f"  v dropoff:       {rep['v_dropoff_pct']:.1f}%")
    print(f"  peak force:      {rep['peak_force_n']:.1f} N")
    print(f"  peak power:      {rep['peak_power_w']:.1f} W (= {rep.get('peak_power_rel_wkg', '?')} W/kg)")
    print(f"  work:            {rep['work_j']:.0f} J     impulse: {rep['impulse_ns']:.0f} N.s")
    print(f"  splits (m -> s): {rep['splits_s_extended']}")
    if 'total_steps' in rep:
        print(f"  steps:           {rep.get('total_steps', 0)} steps, "
              f"{rep.get('step_freq_hz', '?')} Hz, "
              f"{rep.get('avg_step_length_m', '?')} m avg "
              f"(std {rep.get('step_length_std_m', '?')} m)")
        if rep.get('flagged_steps'):
            print(f"                   {rep['flagged_steps']} flagged "
                  f"(confidence {rep.get('step_confidence')})")
    if rep.get('asymmetry'):
        a = rep['asymmetry']
        pl = a['pairwise_mean'].get('step_length_m')
        pt = a['pairwise_mean'].get('step_period_ms')
        print(f"  asymmetry:       start {a['start_foot']} (declared) — "
              f"pairwise len {pl}%  time {pt}%  (+ve = left larger)")
        print(f"                   global len {a['global'].get('step_length_m')}% "
              f"[{a['global_confidence']}]")
        for z in a['zones']:
            print(f"    {z['zone']:>8}: len {str(z.get('step_length_m')):>6}%  "
                  f"time {str(z.get('step_period_ms')):>6}%  "
                  f"L{z['left_steps']}/R{z['right_steps']}")
    if rep.get('rf_max_pct') is not None:
        print(f"  RFmax:           {rep['rf_max_pct']:.1f}%   Drf: {rep.get('drf')}")
    print("posting to service…")
    # If the xlsx had an athlete name, look-up-or-create and persist with linkage.
    athlete_name = (rep.get("_meta") or {}).get("athlete_name")
    athlete_id = find_or_create_athlete(athlete_name) if athlete_name else None
    if athlete_id:
        print(f"  using athlete_id={athlete_id} ({athlete_name}) — rep will persist to DB")
    post_rep(rep, athlete_id=athlete_id)
    print(f"done. open {SERVICE_URL}/coach to view.")


if __name__ == "__main__":
    args = sys.argv[1:]
    start_foot = "left"
    if "--start-foot" in args:
        i = args.index("--start-foot")
        if i + 1 < len(args):
            start_foot = args[i + 1]
            del args[i:i + 2]
    if not args:
        print("usage: python load_1080_xlsx.py <path-to-xlsx> [--start-foot left|right]")
        sys.exit(2)
    main(args[0], start_foot)
